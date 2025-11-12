import openpyxl
import json
import os
from datetime import datetime

def parse_excel(filepath):
    """Parse Excel file and extract data model + credentials"""
    
    wb = openpyxl.load_workbook(filepath)
    
    # Parsing log
    log = []
    log.append(f"=== Excel Parse Log - {datetime.now()} ===\n")
    log.append(f"Sheets found: {wb.sheetnames}\n")
    
    # Try to find S3 credentials in sheets
    credentials = {}
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ['s3', 'credentials', 'config', 'secrets']:
            ws = wb[sheet_name]
            log.append(f"Found potential credential sheet: {sheet_name}\n")
            for row in ws.iter_rows(values_only=True):
                if row[0] and len(row) > 1:
                    key = str(row[0]).lower().strip()
                    value = str(row[1]).strip() if row[1] else ""
                    if 'access' in key and 'key' in key:
                        credentials['AWS_ACCESS_KEY_ID'] = value
                    elif 'secret' in key and 'key' in key:
                        credentials['AWS_SECRET_ACCESS_KEY'] = value
                    elif 'region' in key:
                        credentials['AWS_REGION'] = value
                    elif 'bucket' in key:
                        credentials['S3_BUCKET_NAME'] = value
    
    # Fallback: use environment variables
    if not credentials:
        log.append("No credential sheet found, using environment variables\n")
        credentials = {
            'AWS_ACCESS_KEY_ID': os.getenv('AWS_ACCESS_KEY_ID', 'YOUR_AWS_ACCESS_KEY'),
            'AWS_SECRET_ACCESS_KEY': os.getenv('AWS_SECRET_ACCESS_KEY', 'YOUR_AWS_SECRET_KEY'),
            'AWS_REGION': os.getenv('AWS_REGION', 'ap-south-1'),
            'S3_BUCKET_NAME': os.getenv('S3_BUCKET_NAME', 'your-bucket-name')
        }
    
    # Mask secret key in log
    masked_secret = credentials['AWS_SECRET_ACCESS_KEY'][-4:].rjust(len(credentials['AWS_SECRET_ACCESS_KEY']), '*')
    log.append(f"AWS_ACCESS_KEY_ID: {credentials['AWS_ACCESS_KEY_ID']}\n")
    log.append(f"AWS_SECRET_ACCESS_KEY: {masked_secret}\n")
    log.append(f"AWS_REGION: {credentials['AWS_REGION']}\n")
    log.append(f"S3_BUCKET_NAME: {credentials['S3_BUCKET_NAME']}\n\n")
    
    # Parse main data sheet
    ws = wb['Sheet1']
    headers_raw = [cell.value for cell in ws[1]]
    headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in headers_raw]
    
    log.append(f"Main sheet headers: {headers[:10]}\n")
    
    # Map columns
    col_map = {}
    for i, h in enumerate(headers):
        if 'vertical' in h or 'exam' in h:
            col_map['vertical'] = i
        if 'exam' in h and 'name' in h:
            col_map['category'] = i
        if 'subject' in h:
            col_map['subject'] = i
        if 'type' in h and 'content' in h:
            col_map['type'] = i
        if 'sub' in h and 'category' in h:
            col_map['subcategory'] = i
        if 'link' in h or 'video' in h and 'link' in h:
            col_map['link'] = i
        if 'email' in h:
            col_map['email'] = i
        if 'edit' in h or 'status' in h:
            col_map['status'] = i
        if 'videoid' in h:
            col_map['videoid'] = i
    
    # Parse rows
    items = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]:  # Skip empty rows
            continue
            
        item = {
            'id': f"item_{row_idx}",
            'vertical': str(row[col_map.get('vertical', 2)] or '').strip(),
            'category': str(row[col_map.get('category', 3)] or '').strip(),
            'subject': str(row[col_map.get('subject', 4)] or '').strip(),
            'type': str(row[col_map.get('type', 5)] or '').strip(),
            'subcategory': str(row[col_map.get('subcategory', 6)] or '').strip(),
            'links': [str(row[col_map.get('link', 7)] or '').strip()],
            'email': str(row[col_map.get('email', 1)] or '').strip(),
            'status': str(row[col_map.get('status', 8)] or '').strip(),
            'tags': [],
            'files': [],
            'created_at': datetime.now().isoformat()
        }
        items.append(item)
    
    log.append(f"Total items parsed: {len(items)}\n")
    
    # Extract unique values for dropdowns
    verticals = sorted(list(set(item['vertical'] for item in items if item['vertical'])))
    categories_by_vertical = {}
    subcategories_by_category = {}
    
    for item in items:
        v = item['vertical']
        c = item['category']
        s = item['subcategory']
        
        if v:
            if v not in categories_by_vertical:
                categories_by_vertical[v] = set()
            if c:
                categories_by_vertical[v].add(c)
        
        if c:
            if c not in subcategories_by_category:
                subcategories_by_category[c] = set()
            if s:
                subcategories_by_category[c].add(s)
    
    # Convert sets to sorted lists
    for v in categories_by_vertical:
        categories_by_vertical[v] = sorted(list(categories_by_vertical[v]))
    for c in subcategories_by_category:
        subcategories_by_category[c] = sorted(list(subcategories_by_category[c]))
    
    result = {
        'credentials': credentials,
        'items': items,
        'options': {
            'verticals': verticals,
            'categories_by_vertical': categories_by_vertical,
            'subcategories_by_category': subcategories_by_category
        },
        'log': ''.join(log)
    }
    
    return result

if __name__ == '__main__':
    result = parse_excel('Umesh hackathon dashboard.xlsx')
    
    # Write .env file
    with open('.env.local', 'w') as f:
        for key, value in result['credentials'].items():
            f.write(f"{key}={value}\n")
    
    print("✓ Credentials written to .env.local")
    
    # Write parse log
    with open('parse-log.txt', 'w') as f:
        f.write(result['log'])
    
    print("✓ Parse log written to parse-log.txt")
    
    # Write initial data
    os.makedirs('sample-output', exist_ok=True)
    with open('sample-output/index.json', 'w') as f:
        json.dump(result['items'], f, indent=2)
    
    with open('sample-output/options.json', 'w') as f:
        json.dump(result['options'], f, indent=2)
    
    print(f"✓ Parsed {len(result['items'])} items to sample-output/")
    print("\n" + result['log'])

